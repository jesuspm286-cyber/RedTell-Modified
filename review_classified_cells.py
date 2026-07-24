import argparse
import os

import numpy as np
import pandas as pd
from PIL import Image, ImageDraw
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from skimage.measure import regionprops


parser = argparse.ArgumentParser()
parser.add_argument("--data", required=True)
args = parser.parse_args()

DATA_DIR = args.data

DIAG_PATH = os.path.join(
    DATA_DIR,
    "hierarchical_training_diagnostics.csv",
)

IMAGE_DIR = os.path.join(DATA_DIR, "images")
MASK_DIR = os.path.join(DATA_DIR, "masks")
OUT_DIR = os.path.join(DATA_DIR, "review_cell_crops")
OUT_XLSX = os.path.join(
    DATA_DIR,
    "review_table_with_images.xlsx",
)

os.makedirs(OUT_DIR, exist_ok=True)

diagnostics = pd.read_csv(DIAG_PATH)


review = diagnostics[
    diagnostics["label"].astype(str)
    != diagnostics["final_prediction"].astype(str)
].copy()


# Prioritize errors involving Dead cells, then sort by dead probability.
review["dead_related_error"] = (
    (review["label"] == "Dead")
    | (review["final_prediction"] == "Dead")
)

sort_columns = ["dead_related_error"]
ascending = [False]

if "dead_probability" in review.columns:
    sort_columns.append("dead_probability")
    ascending.append(False)

review = review.sort_values(
    sort_columns,
    ascending=ascending,
)


review["review_label"] = ""
review["review_notes"] = ""

crop_paths = []
# Match masks to images by filename stem, regardless of extension.
mask_lookup = {}

for mask_name in os.listdir(MASK_DIR):
    mask_path = os.path.join(MASK_DIR, mask_name)

    if not os.path.isfile(mask_path):
        continue

    stem = os.path.splitext(mask_name)[0]
    mask_lookup[stem] = mask_path


crop_paths = []

for _, row in review.iterrows():
    fname = os.path.basename(str(row["image"]))
    image_stem = os.path.splitext(fname)[0]
    cell_id = int(row["cell_id"])

    # First try the image path stored in the diagnostics file.
    stored_image_path = str(row["image"])

    if os.path.exists(stored_image_path):
        img_path = stored_image_path
    else:
        img_path = os.path.join(IMAGE_DIR, fname)

    mask_path = mask_lookup.get(image_stem)

    if not os.path.exists(img_path):
        print("Image not found:", img_path)
        crop_paths.append("")
        continue

    if mask_path is None or not os.path.exists(mask_path):
        print("Mask not found for:", fname)
        crop_paths.append("")
        continue

    img = Image.open(img_path).convert("RGB")
    mask = np.array(Image.open(mask_path))

    if mask.ndim == 3:
        mask = mask[:, :, 0]

    cell_mask = mask == cell_id

    if cell_mask.sum() == 0:
        print(
            "Cell ID not found in mask:",
            fname,
            "cell",
            cell_id,
        )
        crop_paths.append("")
        continue

    props = regionprops(
        cell_mask.astype(np.uint8)
    )[0]

    minr, minc, maxr, maxc = props.bbox

    pad = 25

    minr = max(minr - pad, 0)
    minc = max(minc - pad, 0)
    maxr = min(maxr + pad, img.height)
    maxc = min(maxc + pad, img.width)

    crop = img.crop(
        (minc, minr, maxc, maxr)
    )

    draw = ImageDraw.Draw(crop)

    x0 = props.bbox[1] - minc
    y0 = props.bbox[0] - minr
    x1 = props.bbox[3] - minc
    y1 = props.bbox[2] - minr

    draw.rectangle(
        [(x0, y0), (x1, y1)],
        outline="red",
        width=2,
    )

    crop = crop.resize((120, 120))

    crop_name = (
        f"{image_stem}_cell_{cell_id}.png"
    )

    crop_path = os.path.abspath(
        os.path.join(OUT_DIR, crop_name)
    )

    crop.save(crop_path, format="PNG")
    crop_paths.append(crop_path)


review["cell_crop_path"] = crop_paths

print(
    "Crops successfully created:",
    sum(bool(path) for path in crop_paths),
    "of",
    len(crop_paths),
)


review["cell_crop_path"] = crop_paths


cols = [
    "image",
    "cell_id",

    # Original four-class result
    "label",
    "final_prediction",

    # Hierarchical diagnostics
    # "stage1_dead_prediction",
    "dead_probability",
    # "borderline_ghost",
    # "stage2_morphology_prediction",
    "morph_confidence",
    # "dead_missed",
    # "dead_related_error",

    # Manual review columns
    "review_label",
    # "review_notes",
    "cell_crop_path",
]

cols = [
    column
    for column in cols
    if column in review.columns
]

review = review[cols]


# Create Excel workbook.
wb = Workbook()
ws = wb.active
ws.title = "Review"

for row in dataframe_to_rows(
    review,
    index=False,
    header=True,
):
    ws.append(row)


# Header formatting.
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(
        horizontal="center"
    )

ws.freeze_panes = "A2"


# Add cell images.
image_col = ws.max_column + 1

ws.cell(
    row=1,
    column=image_col,
).value = "cell_image"

ws.cell(
    row=1,
    column=image_col,
).font = Font(bold=True)


excel_images = []

for excel_row, crop_path in enumerate(
    review["cell_crop_path"],
    start=2,
):
    ws.row_dimensions[excel_row].height = 95

    if isinstance(crop_path, str) and os.path.isfile(crop_path):
        xl_img = XLImage(crop_path)
        xl_img.width = 90
        xl_img.height = 90

        excel_images.append(xl_img)

        location = ws.cell(
            row=excel_row,
            column=image_col,
        ).coordinate

        ws.add_image(xl_img, location)


# Set readable column widths.
for column_cells in ws.columns:
    column_letter = column_cells[0].column_letter
    header = str(column_cells[0].value)

    if header == "image":
        width = 45
    elif header in {
        "review_notes",
        "cell_crop_path",
    }:
        width = 45
    elif header in {
        "label",
        "final_prediction",
        "stage1_dead_prediction",
        "stage2_morphology_prediction",
        "review_label",
    }:
        width = 22
    else:
        width = 18

    ws.column_dimensions[
        column_letter
    ].width = width


ws.column_dimensions[
    ws.cell(
        row=1,
        column=image_col,
    ).column_letter
].width = 18


wb.save(OUT_XLSX)

print("Confused cells:", len(review))
print("Saved Excel review table to:", OUT_XLSX)
print("Saved crops to:", OUT_DIR)
